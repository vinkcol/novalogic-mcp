"""
ETL 03: Datasets locales (courier + banking) → simora_v2 (novalogic_mcp)
=========================================================================
Fuentes:
  - datasets/courier/domiflash/{anio}/{anio}-{mes}.xlsx  → fact_courier_reports
  - datasets/banking/52500011739/{anio}/52500011739_*.xlsx → fact_bank_transactions

Uso:
  python 03_onedrive_to_simora_v2.py [--source courier|banking|all] [--dry-run]

Resultado: JSON a stdout con totales, logs a stderr.
"""

import os
import sys
import json
import argparse
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Añadir utils/ al path para importar simora_db
sys.path.insert(0, str(Path(__file__).parent.parent / "utils"))
import simora_db

# ─── CONFIG ──────────────────────────────────────────────────────────────────

COMPANY_SLUG = "simora"
BATCH_SIZE   = 200

SCRIPT_DIR   = Path(__file__).parent
DATASETS_DIR = SCRIPT_DIR.parent.parent / "datasets"
COURIER_DIR  = DATASETS_DIR / "courier" / "domiflash"
BANKING_DIR  = DATASETS_DIR / "banking" / "52500011739"

# ─── COURIER PARSING ─────────────────────────────────────────────────────────

def parse_courier_date(val, year: int, month: int) -> str | None:
    """Convierte 'DD/MM/YYYY' o 'DD/MM/YY' a ISO date string."""
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return date(year, month, 1).isoformat()


def to_decimal(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None
    s = str(val).replace(",", "").replace(" ", "").strip()
    try:
        return float(s) if s else None
    except ValueError:
        return None


def to_cop(raw) -> float | None:
    """
    Archivos 2024+ traen valores en miles de COP (7 = 7.000 COP).
    Archivos 2023 (JHW) ya traen COP completo (8500 = 8.500 COP).
    Heuristica: si el valor > 100 ya esta en COP; si <= 100, multiplicar x 1000.
    """
    v = to_decimal(raw)
    if v is None:
        return None
    return v if v > 100 else round(v * 1000, 2)


def load_courier_file(xlsx_path: Path, run_id: str, mcp_conn) -> tuple[int, int, int, int]:
    source_file = xlsx_path.name
    m = re.match(r"(\d{4})-(\d{2})\.xlsx", source_file)
    if not m:
        print(f"  SKIP: nombre inesperado {source_file}", file=sys.stderr)
        return 0, 0, 0, 0

    year, month = int(m.group(1)), int(m.group(2))

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    rows_out = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and any(str(c).strip().lower() == "factura" for c in row if c):
                header_found = True
            continue

        if not row or all(c is None for c in row):
            break

        fecha_raw, factura, valor, flete, destino, visita, novedad, cobro, cliente, descripcion = (
            (row + (None,) * 10)[:10]
        )

        guide = str(factura).strip() if factura else None
        if not guide or guide.lower() in ("factura", "total", ""):
            continue

        report_date = parse_courier_date(fecha_raw, year, month)
        if not report_date:
            report_date = date(year, month, 1).isoformat()

        rows_out.append({
            "source_file":    source_file,
            "report_date":    report_date,
            "guide_number":   guide,
            "declared_value": to_cop(valor),
            "shipping_cost":  to_cop(flete),
            "destination":    str(destino).strip() if destino else None,
            "visit_number":   str(visita).strip() if visita else None,
            "status":         str(novedad).strip() if novedad else None,
            "cash_collected": to_cop(cobro),
            "customer_name":  str(cliente).strip() if cliente else None,
            "description":    str(descripcion).strip() if descripcion else None,
            "etl_run_id":     run_id,
        })

    total_inserted = total_updated = 0
    for i in range(0, len(rows_out), BATCH_SIZE):
        batch = rows_out[i:i + BATCH_SIZE]
        ins, upd = simora_db.upsert_courier_reports(mcp_conn, batch)
        total_inserted += ins
        total_updated  += upd

    return len(rows_out), total_inserted, total_updated, 0


# ─── BANKING PARSING ─────────────────────────────────────────────────────────

def parse_bank_date(val, year: int) -> str | None:
    if not val:
        return None
    s = str(val).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return None
    try:
        return datetime.strptime(s, "%Y/%m/%d").date().isoformat()
    except ValueError:
        pass
    return None


def parse_amount(val) -> float | None:
    if val is None:
        return None
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def load_banking_file(xlsx_path: Path, run_id: str, mcp_conn) -> tuple[int, int, int, int]:
    source_file = xlsx_path.name
    m = re.search(r"([A-Za-z]{3})(\d{4})", source_file)
    if not m:
        print(f"  SKIP: no se pudo extraer anio/mes de {source_file}", file=sys.stderr)
        return 0, 0, 0, 0

    year           = int(m.group(2))
    account_number = "52500011739"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    rows_out       = []
    in_transactions = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        first_cell = str(row[0]).strip() if row[0] is not None else ""

        if not in_transactions:
            if first_cell.upper() == "FECHA":
                in_transactions = True
            continue

        if not row[0]:
            break

        fecha_raw, descripcion, sucursal, dcto, valor_raw, saldo_raw = (
            (row + (None,) * 6)[:6]
        )

        txn_date = parse_bank_date(fecha_raw, year)
        if not txn_date:
            continue

        amount = parse_amount(valor_raw)
        if amount is None:
            continue

        rows_out.append({
            "source_file":      source_file,
            "account_number":   account_number,
            "transaction_date": txn_date,
            "description":      str(descripcion).strip() if descripcion else None,
            "branch":           str(sucursal).strip() if sucursal else None,
            "document":         str(dcto).strip() if dcto else None,
            "amount":           amount,
            "balance":          parse_amount(saldo_raw),
            "etl_run_id":       run_id,
        })

    total_inserted = total_updated = 0
    for i in range(0, len(rows_out), BATCH_SIZE):
        batch = rows_out[i:i + BATCH_SIZE]
        ins, upd = simora_db.upsert_bank_transactions(mcp_conn, batch)
        total_inserted += ins
        total_updated  += upd

    return len(rows_out), total_inserted, total_updated, 0


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run_courier(dry_run: bool) -> dict:
    files = sorted(COURIER_DIR.glob("**/*.xlsx"))
    print(f"[courier] {len(files)} archivos encontrados", file=sys.stderr)

    if dry_run:
        for f in files:
            print(f"  {f.relative_to(DATASETS_DIR)}", file=sys.stderr)
        return {"files": len(files), "dry_run": True}

    mcp_conn = simora_db.get_conn()
    run_id   = simora_db.etl_start(mcp_conn, COMPANY_SLUG, "onedrive", "fact_courier_reports")
    total_p  = total_i = total_u = total_f = 0

    for xlsx in files:
        print(f"  -> {xlsx.name}", file=sys.stderr)
        p, i, u, f = load_courier_file(xlsx, run_id, mcp_conn)
        total_p += p; total_i += i; total_u += u; total_f += f
        print(f"     {p} filas | {i} ins | {u} upd", file=sys.stderr)

    simora_db.etl_finish(mcp_conn, run_id, total_p, total_i, total_u, total_f)
    mcp_conn.close()
    return {"run_id": run_id, "files": len(files),
            "processed": total_p, "inserted": total_i, "updated": total_u, "failed": total_f}


def run_banking(dry_run: bool) -> dict:
    files = sorted(BANKING_DIR.glob("**/*.xlsx"))
    print(f"[banking] {len(files)} archivos encontrados", file=sys.stderr)

    if dry_run:
        for f in files:
            print(f"  {f.relative_to(DATASETS_DIR)}", file=sys.stderr)
        return {"files": len(files), "dry_run": True}

    mcp_conn = simora_db.get_conn()
    run_id   = simora_db.etl_start(mcp_conn, COMPANY_SLUG, "onedrive", "fact_bank_transactions")
    total_p  = total_i = total_u = total_f = 0

    for xlsx in files:
        print(f"  -> {xlsx.name}", file=sys.stderr)
        p, i, u, f = load_banking_file(xlsx, run_id, mcp_conn)
        total_p += p; total_i += i; total_u += u; total_f += f
        print(f"     {p} filas | {i} ins | {u} upd", file=sys.stderr)

    simora_db.etl_finish(mcp_conn, run_id, total_p, total_i, total_u, total_f)
    mcp_conn.close()
    return {"run_id": run_id, "files": len(files),
            "processed": total_p, "inserted": total_i, "updated": total_u, "failed": total_f}


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", choices=["courier", "banking", "all"], default="all")
    parser.add_argument("--dry-run", action="store_true", help="Solo listar archivos, no cargar")
    args = parser.parse_args()

    results = {}

    if args.source in ("courier", "all"):
        results["courier"] = run_courier(args.dry_run)

    if args.source in ("banking", "all"):
        results["banking"] = run_banking(args.dry_run)

    print(json.dumps(results, indent=2))
